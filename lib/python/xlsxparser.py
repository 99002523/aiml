def xlsx(action):
    from openpyxl import load_workbook

    workbook = load_workbook(filename=r"D:\aiml\excelsheet\test_case.xlsx")
    sheet=workbook.active
    for row in sheet.iter_rows(min_row=1,max_row=1):
        for cell in row:
            if cell.value == "testcase":
                col1=cell.column
                row1=cell.row
                print(row1)
                print(col1)
            if cell.value == "keywords":
                col2=cell.column
                row2=cell.row
                print(row2)
                print(col2)

    val=None
    for i in range(1,col2+1):
        #print(sheet.cell(row=i,column=col1).value)
        if sheet.cell(row=i,column=col1).value == action:
            print(action)
            for j in range(1,col2+1):
                val=sheet.cell(row=i,column=col2).value
                val=val.split("\n")
                break
    return val
