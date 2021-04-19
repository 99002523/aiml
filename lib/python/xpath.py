import openpyxl

def xpath(action):
    from openpyxl import load_workbook
    col1=col2=col3=row1=row2=row3=0
    workbook = load_workbook(filename=r"D:\aiml\variables\coordinates.xlsx")
    sheet=workbook.active
    for row in sheet.iter_rows(min_row=1,max_row=1):
        for cell in row:
            if cell.value == "VARIABLE NAME":
                col1=cell.column
                row1=cell.row
                #print(row1)
                #print(col1)
            if cell.value == "COORDINATES":
                col2=cell.column
                row2=cell.row
                #print(row2)
                #print(col2)
            if cell.value == "XPATH":
                col3=cell.column
                row3=cell.row
                #print(row3)
                #print(col3)
    #print(action)
    rows=sheet.max_row
    #print(rows)
    val=[None,None]
    for i in range(1,rows+1):
        #print(sheet.cell(row=i,column=col1).value)
        if sheet.cell(row=i,column=col1).value == action:
            #print(action)
            for j in range(1,rows+1):
                val=sheet.cell(row=i,column=col2).value
                if val!=None:
                    val=val.split("\n")
                break
            for j in range(1,rows+1):
                if col3==0:
                    break
                val3=sheet.cell(row=i,column=col3).value
                if val3!=None:
                    val.append(val3)
                break
    return val

#uncomment this to check how this function is working
#a=xpath("playbutton")
#print(a)
