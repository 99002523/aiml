import re
import openpyxl
import os

import matplotlib.pyplot as plt
import cv2
import easyocr
from pylab import rcParams
from IPython.display import Image

def timextractor():
    rcParams['figure.figsize'] = 8, 16
    reader = easyocr.Reader(['en', 'hi'])
    os.chdir('/aiml/screenshot')
    cwd=os.getcwd()
    file_name =cwd+r"\JioCinema.png"
    Image(file_name)
    output = reader.readtext(file_name)
    print(output)
    ls=[lis[-2] for lis in output]
    #ls=output
    a=[]
    time=[]
    for item in ls:
        if "/" in item:    
            time=item.split('/')
            for i in time:
               x= re.findall("(?:(?:([01]?\d|2[0-3]):)?([0-5]?\d):)([0-5]?\d)$",i)
               if x!=[]:
                  a.append(x[0])
        else:     
            x= re.findall("(?:(?:([01]?\d|2[0-3]):)?([0-5]?\d):)([0-5]?\d)$",item)
            if x!=[]:
               a.append(x[0])
    wb= openpyxl.Workbook()
    sheet=wb.active
    i=1
    j=1
    c1=sheet.cell(row=i,column=j)
    c2=sheet.cell(row=i,column=j+1)
    c3=sheet.cell(row=i,column=j+2)
    c1.value="HOURS"
    c2.value="MINUTES"
    c3.value="SECONDS"
    i=2
    j=1
    for z in range(i,len(a)+2):
        c1=sheet.cell(row=z,column=j)
        c2=sheet.cell(row=z,column=j+1)
        c3=sheet.cell(row=z,column=j+2)
        c1.value=a[z-2][0]
        c2.value=a[z-2][1]
        c3.value=a[z-2][2]
    os.chdir('/aiml/variables')
    cwd=os.getcwd()
    wb.save(cwd+"\\timestamp.xlsx")
    #print(a)

#timextractor(lis)





def timechecker():
    from openpyxl import load_workbook
    col1=col2=col3=row1=row2=row3=0
    os.chdir('/aiml/variables')
    cwd=os.getcwd()
    workbook = load_workbook(filename=cwd+r"\timestamp.xlsx")
    sheet=workbook.active
    for row in sheet.iter_rows(min_row=1,max_row=1):
        for cell in row:
            if cell.value == "HOURS":
                col1=cell.column
                row1=cell.row
                #print(row1)
                #print(col1)
            if cell.value == "MINUTES":
                col2=cell.column
                row2=cell.row
                #print(row2)
                #print(col2)
            if cell.value == "SECONDS":
                col3=cell.column
                row3=cell.row
                #print(row3)
                #print(col3)
    #print(action)
    rows=sheet.max_row
    #print(rows)
    val=[]
    for i in range(2,3):
        #print(sheet.cell(row=i,column=col1).value)
        val1=sheet.cell(row=i,column=col1).value
        if val1 == None:
            val.append("00")
        else:
            val.append(val1)
        val2=sheet.cell(row=i,column=col2).value
        if val2 == None:
            val.append("00")
        else:
            val.append(val2)
        val3=sheet.cell(row=i,column=col3).value
        if val3 == None:
            val.append("00")
        else:
            val.append(val3)
    return val

#uncomment this to check how this function is working
#a=timechecker()
#print(a)



def time_checker_before():
    #lis=["12:13/13:14","1:2:3/2:3:4"]
    timextractor()
    b=timechecker()
    return b
    

def time_checker_after():
    #lis=["12:13/13:14","1:2:3/2:3:4"]
    timextractor()
    a=timechecker()
    return a

def video_playing(before, after):
    flag=0
    #before=time_checker_before()
    #after=time_checker_after()
    for i in range(0,len(before)):
        after_time=int(after[0])*60*60+int(after[1])*60+int(after[2])
        before_time=int(before[0])*60*60+int(before[1])*60+int(before[2])
        if  after_time - before_time > 0:
            flag=1
    if flag:
        return "video is playing"
    else:
        return "video is not playing"
#video_playing()
