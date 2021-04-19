'''
* The following code is an object detection model which uses YOLO V3 learning model type.
* This is a model type inside the imageai.detection library.
* The imageai uses tensorflow in the background to run the algorithm which  is nothing but
  an end-end  AI framework.
* ALGORITHM:- YOLOV3

'''


from imageai.Detection.Custom import CustomObjectDetection
import os
import csv
import openpyxl

def midpoint(z,lis):
        a=((lis[0]+lis[2])/2)
        b=((lis[1]+lis[3])/2)
        #z=z+str(count)
        return [z,a,b]

def detect():
        execution_path = os.getcwd()
        detector = CustomObjectDetection()
        detector.setModelTypeAsYOLOv3()
        detector.setModelPath(r"D:\aiml\lib\python\detection_model-ex-007--loss-0008.640.h5")
        detector.setJsonPath(r"D:\aiml\lib\python\detection_config.json")
        detector.loadModel()
        count=0
        detections = detector.detectObjectsFromImage(input_image=r"D:\\aiml\\screenshot\\input.png", output_image_path=r"D:\\aiml\\screenshot\\output.png")
        wb= openpyxl.Workbook()
        sheet=wb.active
        i=1
        j=1
        c1=sheet.cell(row=i,column=j)
        c2=sheet.cell(row=i,column=j+1)
        c1.value="VARIABLE NAME"
        c2.value="COORDINATES"
        i=2
        j=1
        for eachObject in detections:
            for i in range(i,25):
                print(eachObject["name"] , " : " , eachObject["percentage_probability"] , ":" , eachObject["box_points"])
                c1=sheet.cell(row=i,column=j)
                c2=sheet.cell(row=i,column=j+1)
                object_value=midpoint(eachObject["name"],eachObject["box_points"])
                c1.value=object_value[0]
                tom=str(object_value[1])+str(",")+str(object_value[2])
                c2.value=tom
                i+=1
                print(midpoint(eachObject["name"],eachObject["box_points"]))
                break
                print(midpoint(eachObject["name"],eachObject["box_points"]))
                count=0
        wb.save("D:\\aiml\\variables\\coordinates.xlsx")
detect()
