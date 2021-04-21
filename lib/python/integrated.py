'''
* The following code is an object detection model which uses YOLO V3 learning model type.
* This is a model type inside the imageai.detection library.
* The imageai uses tensorflow in the background to run the algorithm which  is nothing but
  an end-end  AI framework.
* ALGORITHM:- YOLOV3

'''


from imageai.Detection.Custom import CustomObjectDetection
import os
import csv
import openpyxl

def midpoint(z,lis):
        a=((lis[0]+lis[2])/2)
        b=((lis[1]+lis[3])/2)+100
        #z=z+str(count)
        return [z,a,b]

def detect():
        execution_path = os.getcwd()
        detector = CustomObjectDetection()
        detector.setModelTypeAsYOLOv3()
        detector.setModelPath(r"D:\aiml\lib\python\detection_model-ex-006--loss-0009.732.h5")
        detector.setJsonPath(r"D:\aiml\lib\python\detection_config.json")
        detector.loadModel()
        count=0
        detections = detector.detectObjectsFromImage(input_image=r"D:\\aiml\\screenshot\\input.png", output_image_path=r"D:\\aiml\\screenshot\\output.png")
        wb= openpyxl.Workbook()
        sheet=wb.active
        i=1
        j=1
        c1=sheet.cell(row=i,column=j)
        c2=sheet.cell(row=i,column=j+1)
        c1.value="VARIABLE NAME"
        c2.value="COORDINATES"
        i=2
        j=1
        for eachObject in detections:
            for i in range(i,25):
                print(eachObject["name"] , " : " , eachObject["percentage_probability"] , ":" , eachObject["box_points"])
                c1=sheet.cell(row=i,column=j)
                c2=sheet.cell(row=i,column=j+1)
                object_value=midpoint(eachObject["name"],eachObject["box_points"])
                c1.value=object_value[0]
                tom=str(object_value[1])+str(",")+str(object_value[2])
                c2.value=tom
                i+=1
                print(midpoint(eachObject["name"],eachObject["box_points"]))
                break
                print(midpoint(eachObject["name"],eachObject["box_points"]))
                count=0
        wb.save("D:\\aiml\\variables\\coordinates.xlsx")

#detect()


"""
Qxf2 Services: Utility script to generate XPaths for the given URL
* Take the input URL from the user
* Parse the HTML content using beautifilsoup
* Find all Input and Button tags
* Guess the XPaths
* Generate Variable names for the xpaths
* To run the script in Gitbash use command 'python -u utils/xpath_util.py'
"""
from selenium import webdriver
from bs4 import BeautifulSoup
import re
import openpyxl
wb=openpyxl.Workbook()
sheet=wb.active
class Xpath_Util:
    "Class to generate the xpaths"
    def __init__(self):
        "Initialize the required variables"
        self.elements = None
        self.guessable_elements = ['input','button']
        self.known_attribute_list = ['id','name','placeholder','value','title','type','class']
        self.variable_names = []
        self.button_text_lists = []
        self.language_counter = 1
        
    def generate_xpath(self,soup,driver,xpath_obj):
        i=1
        j=1
        c1=sheet.cell(row=1,column=1)
        c1.value="VARIABLE NAME"
        c2=sheet.cell(row=1,column=2)
        c2.value="XPATH"
        "generate the xpath and assign the variable names"
        result_flag = False
        for guessable_element in self.guessable_elements:
            self.elements = soup.find_all(guessable_element)
            for element in self.elements:
                if (not element.has_attr("type")) or (element.has_attr("type") and element['type'] != "hidden"):
                    for attr in self.known_attribute_list:
                        if element.has_attr(attr):
                            locator = self.guess_xpath(guessable_element,attr,element)
                            if len(driver.find_elements_by_xpath(locator))==1:
                                result_flag = True
                                variable_name = self.get_variable_names(element)
                                # checking for the unique variable names
                                if variable_name != '' and variable_name not in self.variable_names:
                                    self.variable_names.append(variable_name)
                                    i=i+1                               
                                    c1=sheet.cell(row=i,column=j) 
                                    c1.value="%s_%s"%(guessable_element, variable_name.encode('utf-8').decode('latin-1'))                                                
                                    j=j+1
                                    c2=sheet.cell(row=i,column=j)
                                    c2.value="%s"%(locator.encode('utf-8').decode('latin-1'))
                                    j=j-1
                                    print ("%s_%s = %s"%(guessable_element, variable_name.encode('utf-8').decode('latin-1'), locator.encode('utf-8').decode('latin-1')))
                                    break
                                else:
                                    i=i+1
                                    c1=sheet.cell(row=i,column=j)      
                                    c1.value="Couldn't generate appropriate variable name for this xpath"
                                    j=j+1
                                    c2=sheet.cell(row=i,column=j)
                                    c2.value=locator.encode('utf-8').decode('latin-1')
                                    j=j-1
                                    print (locator.encode('utf-8').decode('latin-1') + "----> Couldn't generate appropriate variable name for this xpath")
                        elif guessable_element == 'button' and element.getText():
                            button_text = element.getText()
                            if element.getText() == button_text.strip():
                                locator = xpath_obj.guess_xpath_button(guessable_element,"text()",element.getText())
                            else:
                                locator = xpath_obj.guess_xpath_using_contains(guessable_element,"text()",button_text.strip())
                            if len(driver.find_elements_by_xpath(locator))==1:
                                result_flag = True
                                #Check for utf-8 characters in the button_text
                                matches = re.search(r"[^\x00-\x7F]",button_text)
                                if button_text.lower() not in self.button_text_lists:
                                    self.button_text_lists.append(button_text.lower())
                                    if not matches:
                                        # Striping and replacing characters before printing the variable name
                                        i=i+1
                                        c1=sheet.cell(row=i,column=j)
                                        c1.value="%s_%s"%(guessable_element,button_text.strip().strip("!?.").encode('utf-8').decode('latin-1').lower().replace(" + ","_").replace(" & ","_").replace(" ","_"))
                                        j=j+1
                                        c2=sheet.cell(row=i,column=j)
                                        c2.value="%s"%(locator.encode('utf-8').decode('latin-1'))
                                        j=j-1
                                        print ("%s_%s = %s"%(guessable_element,button_text.strip().strip("!?.").encode('utf-8').decode('latin-1').lower().replace(" + ","_").replace(" & ","_").replace(" ","_"), locator.encode('utf-8').decode('latin-1')))
                                    else:
                                        # printing the variable name with utf-8 characters along with language counter
                                        i=i+1
                                        c1=sheet.cell(row=i,column=j)
                                        c1.value="%s_%s_%s"%(guessable_element,"foreign_language",self.language_counter) + "---> Foreign language found, please change the variable name appropriately"
                                        j=j+1
                                        c2=sheet.cell(row=i,column=j)
                                        c2.value="%s"%(locator.encode('utf-8').decode('latin-1'))
                                        j=j-1
                                        print ("%s_%s_%s = %s"%(guessable_element,"foreign_language",self.language_counter, locator.encode('utf-8').decode('latin-1')) + "---> Foreign language found, please change the variable name appropriately")
                                        self.language_counter +=1
                                else:
                                    # if the variable name is already taken
                                    i=i+1
                                    c1=sheet.cell(row=i,column=j)
                                    c1.value="Couldn't generate appropriate variable name for this xpath"
                                    j=j+1
                                    c2=sheet.cell(row=i,column=j)
                                    c2.value=locator.encode('utf-8').decode('latin-1')
                                    j=j-1
                                    print (locator.encode('utf-8').decode('latin-1') + "----> Couldn't generate appropriate variable name for this xpath")
                                break
                        elif not guessable_element in self.guessable_elements:
                            i=i+1
                            c1=sheet.cell(row=i,column=j)
                            c1.value="We are not supporting this gussable element"           
                            print("We are not supporting this gussable element")   
        wb.save("D:\\aiml\\variables\\xpath_only.xlsx")
        return result_flag
              
            
    def get_variable_names(self,element):
        "generate the variable names for the xpath"
        # condition to check the length of the 'id' attribute and ignore if there are numerics in the 'id' attribute. Also ingnoring id values having "input" and "button" strings.
        if (element.has_attr('id') and len(element['id'])>2) and bool(re.search(r'\d', element['id'])) == False and ("input" not in element['id'].lower() and "button" not in element['id'].lower()):
            self.variable_name = element['id'].strip("_")
        # condition to check if the 'value' attribute exists and not having date and time values in it.
        elif element.has_attr('value') and element['value'] != '' and bool(re.search(r'([\d]{1,}([/-]|\s|[.])?)+(\D+)?([/-]|\s|[.])?[[\d]{1,}',element['value']))== False and bool(re.search(r'\d{1,2}[:]\d{1,2}\s+((am|AM|pm|PM)?)',element['value']))==False:
            # condition to check if the 'type' attribute exists
            # getting the text() value if the 'type' attribute value is in 'radio','submit','checkbox','search'
            # if the text() is not '', getting the getText() value else getting the 'value' attribute
            # for the rest of the type attributes printing the 'type'+'value' attribute values. Doing a check to see if 'value' and 'type' attributes values are matching.
            if (element.has_attr('type')) and (element['type'] in ('radio','submit','checkbox','search')):
                if element.getText() !='':
                    self.variable_name = element['type']+ "_" + element.getText().strip().strip("_.")
                else:
                    self.variable_name = element['type']+ "_" + element['value'].strip("_.")
            else:
                if element['type'].lower() == element['value'].lower():
                    self.variable_name = element['value'].strip("_.")
                else:
                    self.variable_name = element['type']+ "_" + element['value'].strip("_.")
        # condition to check if the "name" attribute exists and if the length of "name" attribute is more than 2 printing variable name
        elif element.has_attr('name') and len(element['name'])>2:
            self.variable_name = element['name'].strip("_")
        # condition to check if the "placeholder" attribute exists and is not having any numerics in it.
        elif element.has_attr('placeholder') and bool(re.search(r'\d', element['placeholder'])) == False:
            self.variable_name = element['placeholder']
	# condition to check if the "type" attribute exists and not in text','radio','button','checkbox','search'
        # and printing the variable name
        elif (element.has_attr('type')) and (element['type'] not in ('text','button','radio','checkbox','search')):
            self.variable_name = element['type']
        # condition to check if the "title" attribute exists
        elif element.has_attr('title'):
            self.variable_name = element['title']
        # condition to check if the "role" attribute exists
        elif element.has_attr('role') and element['role']!="button":
            self.variable_name = element['role']
        else:
            self.variable_name = ''
        return self.variable_name.lower().replace("+/- ","").replace("| ","").replace(" / ","_"). \
        replace("/","_").replace(" - ","_").replace(" ","_").replace("&","").replace("-","_"). \
        replace("[","_").replace("]","").replace(",","").replace("__","_").replace(".com","").strip("_")
    def guess_xpath(self,tag,attr,element):
        "Guess the xpath based on the tag,attr,element[attr]"
        #Class attribute returned as a unicodeded list, so removing 'u from the list and joining back
        if type(element[attr]) is list:
            element[attr] = [i.encode('utf-8').decode('latin-1') for i in element[attr]]
            element[attr] = ' '.join(element[attr])
        self.xpath = "//%s[@%s='%s']"%(tag,attr,element[attr])
        return self.xpath
    def guess_xpath_button(self,tag,attr,element):
        "Guess the xpath for button tag"
        self.button_xpath = "//%s[%s='%s']"%(tag,attr,element)
        return self.button_xpath
    def guess_xpath_using_contains(self,tag,attr,element):
        "Guess the xpath using contains function"
        self.button_contains_xpath = "//%s[contains(%s,'%s')]"%(tag,attr,element)
        return self.button_contains_xpath
#-------START OF SCRIPT--------
def xextract(driver,url,session_id):

    #print ("Start of %s"%__file__)
    #driver = webdriver.Chrome(r"D:\aiml\driver\chromedriver.exe")
    #driver.get(url)
    #page = driver.execute_script("return document.body.innerHTML").\
    #encode('utf-8').decode('latin-1')
    #soup = BeautifulSoup(page, 'html.parser')
    #if xpath_obj.generate_xpath(soup,driver,xpath_obj) is False:
    #    print ("No XPaths generated for the URL:%s"%url)
    #driver.quit()
    xpath_obj = Xpath_Util()
    driver2 = webdriver.Remote(command_executor=url,desired_capabilities={})
    driver2.session_id = session_id
    page = driver2.execute_script("return document.body.innerHTML").\
    encode('utf-8').decode('latin-1')#returns the inner HTML as a string
    soup = BeautifulSoup(page, 'html.parser')
    #execute generate_xpath
    if xpath_obj.generate_xpath(soup,driver2,xpath_obj) is False:
        print("1")
        print ("No XPaths generated for the URL:%s"%url)
    driver.quit()

#xextract("https://www.netflix.com/in/login")    




from openpyxl import load_workbook

def merger():
    usr=['input_username','input_login_field','username','user','customer_name','email','input_id_userloginid','input_ap_email','input_email_or_msisdn']
    pas=['input_password','pass','passowrd','input_id_password','input_ap_password','input_password']
    btn=['button_sign_in','sign_in','button_signin','input_signinsubmit','button_sign_in','input_submit_sign_in','button_login','input_submit_continue']
    
    workbook1 = load_workbook(filename=r"D:\aiml\variables\coordinates.xlsx")
    sheet2=workbook1.active
    for row in sheet2.iter_rows(min_row=1,max_row=1):
        for cell in row:
            if cell.value == "VARIABLE NAME":
                col3=cell.column
                row3=cell.row
                # print(row1)
                # print(col1)
            if cell.value == "COORDINATES":
                col4=cell.column
                row4=cell.row
                # print(row2)
                # print(col2)
    
    
    workbook = load_workbook(filename=r"D:\\aiml\\variables\\xpath_only.xlsx")
    sheet=workbook.active
    for row in sheet.iter_rows(min_row=1,max_row=1):
        for cell in row:
            if cell.value == "VARIABLE NAME":
                col1=cell.column
                row1=cell.row
                # print(row1)
                # print(col1)
            if cell.value == "XPATH":
                col2=cell.column
                row2=cell.row
                # print(row2)
                # print(col2)

    val=None
    rows=sheet.max_row
    
    for i in range(1,rows):
        #print(sheet.cell(row=i,column=col1).value)
        d=sheet.cell(row=i,column=col1).value
        
        if(d in usr):
            x=sheet.cell(row=i,column=col1+1).value
            print(x)
            rows2=sheet2.max_row
            for j in range(1,rows2+1):
                sc=sheet2.cell(row=row3,column=col4+1)
                sc.value='XPATH'
                d1=sheet2.cell(row=j,column=col3).value
                
                if(d1 == 'usernametextbox'):
                    
                    wr=sheet2.cell(row=j,column=col4+1)
                    wr.value=x
        if(d in pas):
            
            x=sheet.cell(row=i,column=col1+1).value
            print(x)
            rows2=sheet2.max_row
            for j in range(1,rows2+1):
                sc=sheet2.cell(row=row3,column=col4+1)
                sc.value='XPATH'
                d1=sheet2.cell(row=j,column=col3).value
                if(d1 == 'passwordtextbox'):
                    wr=sheet2.cell(row=j,column=col4+1)
                    wr.value=x
                    
        if(d in btn):
            x=sheet.cell(row=i,column=col1+1).value
            print(x)
            rows2=sheet2.max_row
            for j in range(1,rows2+1):
                sc=sheet2.cell(row=row3,column=col4+1)
                sc.value='XPATH'
                d1=sheet2.cell(row=j,column=col3).value
                if( d1=='loginbutton'):
                    wr=sheet2.cell(row=j,column=col4+1)
                    wr.value=x
                    
    workbook1.save("D:\\aiml\\variables\\coordinates.xlsx")
    
    
#merger()


def session_creator(url, cookies):
    from selenium import webdriver
    driver = webdriver.Chrome()
    executor_url = driver.command_executor._url
    session_id = driver.session_id
    driver.get(url)
    #driver.quit()
    xextract(driver,executor_url, session_id)