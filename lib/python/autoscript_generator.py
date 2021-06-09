#importing python modules
import sys
import getopt
import openpyxl
import os
from datetime import datetime

#function to work with xlsx, tha is it reads the input file containing testcase and keywords and return the list containing test cases and keywords
def xlsx(filepath):
    from openpyxl import load_workbook
    workbook = load_workbook(filename=filepath)
    sheet=workbook.active
    for row in sheet.iter_rows(min_row=1,max_row=1):
        for cell in row:
            if cell.value == "title":
                col1=cell.column
                row1=cell.row
                #print(row1)
                #print(col1)
            if cell.value == "keywords":
                col2=cell.column
                row2=cell.row
                #print(row2)
                #print(col2)
            if cell.value == "arguments":
                col3=cell.column
                row3=cell.row
    rows=sheet.max_row
    #val=[]
    result=[]
    for i in range(row1+1,rows+1):
        
        val=[]
        val1=sheet.cell(row=i,column=col1).value
        val.append(val1)
        #for j in range(row1+1,rows+1):
        val2=sheet.cell(row=i,column=col2).value
        val2=val2.split("\n")
        val.append(val2)
        val3=sheet.cell(row=i,column=col3).value
        val3=val3.split("\n")
        val.append(val3)
        #print(val)
        result.append(val)
    #print(result)
    return result

#auto script generator
#print("this is the name of the script",sys.argv[0])
#print("this is file name", sys.argv[1])
#print("url is", sys.argv[2])
if sys.argv[0] != None:
    value=xlsx(sys.argv[1])
    #print(value)
    for element in value:
        filename=element[0]
        testname=element[1]
        variables=element[2]
        if len(testname)!=len(variables):
            #print("error")
            break
        else:
            print("ok")
                      
        try:
            os.chdir('/aiml')
            cwd=os.getcwd()
            os.mkdir(cwd+"/testsuite")
        except OSError as error:
            pass

        try:
            os.chdir('/aiml/testsuite')
            cwd=os.getcwd()
            filepath=cwd+"/{apr}.robot".format(apr=filename)
            f=open(filepath,"w+")
        except:
            print("file exist")

        #script writing
        f=open(filepath,"a")
        f.write("##########################################################\n")
        f.write("#script name : {apr}\n".format(apr=filename))
        f.write("#description : {apr}\n".format(apr=filename))
        f.write("#created by : Auto generated\n")
        now=datetime.now()
        this_time=now.strftime("%Y-%m-%d %H:%M:%S")
        f.write("#created date : {time}\n".format(time=this_time))
        f.write("#reviced by : \n")
        f.write("#copy rights : LTTS\n")
        f.write("#########################################################\n")

        f.write("\n***Settings***\n")
        f.write("Documentation    {apr}\n\n".format(apr=filename))
        f.write("#importing Libraries\n")
        f.write("Resource    ../lib/robot/keywords-coordinates.robot\n")
        f.write("Library    SeleniumLibrary    screenshot_root_directory=screenshot\n")
        f.write("Library    OperatingSystem\n\n")
        
        #testcase
        f.write("\n***Test Cases***\n")
        f.write("{apr}\n".format(apr=filename))
        f.write("    [Setup]    Test Setup\n\n")
        #f.write("    Launch App    {url}\n\n".format(url=variables[0]))
        count_elements=len(testname)
        for i in range(0,count_elements):

            var=variables[i]
            #print(var)
            #print(testname[i])
            result=''
            if var != 'None':
                result = [x.strip() for x in var.split(';')]
                #print(result)
                #print(len(result))

            if len(result) == 0:
                f.write("    ${status}     {key}\n".format(key=testname[i],status="{status}"))
                f.write("    Run Keyword If     \"${status}\"==\"True\"     Log To Console     {key} successful\n".format(key=testname[i],status="{status}"))
                f.write("    ...    ELSE     Fail     {key} failed\n\n".format(key=testname[i]))
            if len(result) == 1:
                f.write("    ${status}     {key}    {param1}\n".format(key=testname[i],status="{status}",param1=result[0]))
                f.write("    Run Keyword If     \"${status}\"==\"True\"     Log To Console     {key} successful\n".format(key=testname[i],status="{status}"))
                f.write("    ...    ELSE     Fail     {key} failed\n\n".format(key=testname[i]))
            if len(result) == 2:
                f.write("    ${status}     {key}    {param1}    {param2}\n".format(key=testname[i],status="{status}",param1=result[0],param2=result[1]))
                f.write("    Run Keyword If     \"${status}\"==\"True\"     Log To Console     {key} successful\n".format(key=testname[i],status="{status}"))
                f.write("    ...    ELSE     Fail     {key} failed\n\n".format(key=testname[i]))
            if len(result) == 3:
                f.write("    ${status}     {key}    {param1}    {param2}    {param3}\n".format(key=testname[i],status="{status}",param1=result[0],param2=result[1],param3=result[2]))
                f.write("    Run Keyword If     \"${status}\"==\"True\"     Log To Console     {key} successful\n".format(key=testname[i],status="{status}"))
                f.write("    ...    ELSE     Fail     {key} failed\n\n".format(key=testname[i]))
        
        f.write("    [Teardown]    Run Keyword If Test Failed    Test Fail Teardown\n")
        f.close()
        print("auto script generated successfully for {testname}\n".format(testname=filename))

