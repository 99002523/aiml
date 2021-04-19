#python code to parse excel sheet containing varaibles, xapth and coordinates
def xpath(action,filepath):
    from openpyxl import load_workbook
    
    workbook = load_workbook(filename=filepath)
    sheet=workbook.active
    for row in sheet.iter_rows(min_row=1,max_row=1):
        for cell in row:
            if cell.value == "variables":
                col1=cell.column
                row1=cell.row
                print(row1)
                print(col1)
            if cell.value == "coordinates":
                col2=cell.column
                row2=cell.row
                print(row2)
                print(col2)
            if cell.value == "xpath":
                col3=cell.column
                row3=cell.row
                print(row3)
                print(col3)
    print(action)
    rows=sheet.max_row
    print(rows)
    val=[]
    result=[]
    for i in range(2,rows+1):
        val=[]
        val1=sheet.cell(row=i,column=col1).value
        print(val1)
        val.append(val1)
        for j in range(1,rows+1):
            val2=sheet.cell(row=i,column=col2).value
            print(val2)
            val.append(val2)
            break
        for j in range(1,rows+1):
            val3=sheet.cell(row=i,column=col3).value
            val.append(val3)
            break
        result.append(val)
    return result



#xls parser that is old excel sheet(microsoft excel 2003) canbe parsed using below code
def xls(action):
    # Reading an excel file using Python
    import xlrd
 
    # Give the location of the file
    loc = r"C:\Users\20130212\eclipse-workspace\aiml\excelsheet\test1.xls"
 
    # To open Workbook
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)

    # reading and storing in var
    count=sheet.nrows
    columns=sheet.ncols
    print(count)
    print(columns)
    i=0
    li=[]
    while i !=count:
        apr=None
        name = sheet.cell_value(i, 0)
        if name == action:
            for j in range(sheet.ncols):
                col=sheet.cell_value(0,j)
                if col == "keywords":
                    apr=sheet.cell_value(i,j)
                    print(apr)
                    li=apr.split("\n")
                    break
        i=i+1
    return li
