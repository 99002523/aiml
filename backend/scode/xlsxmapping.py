from openpyxl import load_workbook

 

def merger():
    usr=['input_username','input_login_field','username','user','customer_name','email','input_id_userloginid','input_ap_email']
    pas=['input_password','pass','passowrd','input_id_password','input_ap_password']
    btn=['button_sign_in','sign_in','button_signin','input_signinsubmit','button_sign_in']
    
    workbook1 = load_workbook(filename=r"D:\backend\xlsx\coordinates.xlsx")
    sheet2=workbook1.active
    for row in sheet2.iter_rows(min_row=1,max_row=1):
        for cell in row:
            if cell.value == "VARIABLE NAME":
                col3=cell.column
                row3=cell.row
                # print(row1)
                # print(col1)
            if cell.value == "COORDINATES":
                col4=cell.column
                row4=cell.row
                # print(row2)
                # print(col2)
    
    
    workbook = load_workbook(filename=r"D:\backend\xlsx\demo.xlsx")
    sheet=workbook.active
    for row in sheet.iter_rows(min_row=1,max_row=1):
        for cell in row:
            if cell.value == "VARIABLE NAME":
                col1=cell.column
                row1=cell.row
                # print(row1)
                # print(col1)
            if cell.value == "XPATH":
                col2=cell.column
                row2=cell.row
                # print(row2)
                # print(col2)

 

    val=None
    rows=sheet.max_row
    
    for i in range(1,rows):
        #print(sheet.cell(row=i,column=col1).value)
        d=sheet.cell(row=i,column=col1).value
        if(d in usr or d in pas or d in btn):
            x=sheet.cell(row=i,column=col1+1).value
            print(x)
            rows2=sheet2.max_row
            for j in range(1,rows2):
                sc=sheet2.cell(row=row3,column=col4+1)
                sc.value='XPATH'
                d1=sheet2.cell(row=i,column=col3).value
                if(d1 == 'usernametextbox'):
                    wr=sheet2.cell(row=i,column=col4+1)
                    wr.value=x
                if(d1 == 'passwordtextbox'):
                    wr=sheet2.cell(row=i,column=col4+1)
                    wr.value=x
                if(d1=='loginbutton'):
                    wr=sheet2.cell(row=i,column=col4+1)
                    wr.value=x
    workbook1.save("D:\\backend\\output\\coordinates.xlsx")
merger()